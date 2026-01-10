from openpyxl import load_workbook
import csv

wb = load_workbook('formula.xlsx', data_only=True)
ws = wb['Misc']

headers = [
    'Car',
    'Height', 'F axle offset', 'R axle offset', 'Wheelbase', 'F track width', 'R track width', 'AV track width',
    'Gears', 'Shift time', 'Weight',
    'Stock CX', 'Stock SX', 'Stock Drag',
    'Stage 1/2 CX', 'Stage 1/2 SX', 'Stage 1/2 Drag',
    'Stage 3/4 CX', 'Stage 3/4 SX', 'Stage 3/4 Drag',
    'Body Position X', 'Body Position Y', 'Body Position Z',
    'Power HP', 'Mass Kg', 'Turbo Press.', 'Curve fall @ RPM', 'Rev limiter', 'Inertia ratio',
    'Engine Position X', 'Engine Position Y', 'Engine Position Z'
]

with open('src/assets/car-data.csv', 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(headers)

    # Data starts at row 4, goes to max_row
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=32, values_only=True):
        # Skip empty rows (first column is car name)
        if row[0] is None or str(row[0]).strip() == '':
            continue

        # Process row values
        processed = []
        for i, val in enumerate(row):
            if val is None:
                processed.append('')
            elif val == 'NA':
                processed.append('')
            elif isinstance(val, float):
                # Clean up floating point precision issues
                if val == int(val):
                    processed.append(int(val))
                else:
                    # Round to reasonable precision
                    processed.append(round(val, 6))
            else:
                processed.append(val)

        writer.writerow(processed)

print(f"Exported {ws.max_row - 3} rows to src/assets/car-data.csv")
